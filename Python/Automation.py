import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook():
     wb = xl.load_workbook('C:\VScode\JavaScript\Python\Xlsx\Transactions2.xlsx')
     sheet = wb['Sheet1']

     
     for row in range(1, sheet.max_row+1):               # sheet.max_row is to get the num of rows in the xlsx and + 1 so that it considers the last row as well
          print(row)
     print()


     for row in range(2, sheet.max_row+1):               # we start from 2 cz we dont want the product id text in the 1st row
          cell = sheet.cell(row,3)                       # (row,3) means that num no row and 3rd cell
          corrected_price = cell.value * 0.1             # we accessed the cell variabel above value here instead of printing and chngd it by multiplying 0.9
          corrected_price_cell = sheet.cell(row,5)       # Here we declared another variable and said we are accessing that rows 4th cell in the sheet
          corrected_price_cell.value = corrected_price   # assigned new value to the cell we just accessed above

          values =  Reference(sheet,                     # select from sheet
                    min_row=2,                           # start from row 2
                    max_row=sheet.max_row,               # end at max_row 
                    min_col=4,                           # we only want col 4 and not all col
                    max_col=4,)
          chart = BarChart()                             # created ref of BarChart
          chart.add_data(values)
          sheet.add_chart(chart,'h2')                    #add chart named 'chart' at cell f2


     wb.save('Transactions3.xlsx')